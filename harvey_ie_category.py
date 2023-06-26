import scrapy
import re
import warnings
import openpyxl
from os import getcwd,path
from datetime import datetime
from bs4  import BeautifulSoup
import os, os.path, time, json, pathlib
from datetime import datetime
import sqlite3 as sql
import pandas as pd


class HarveIeSpider(scrapy.Spider):
    name = "harve_ie"
    start_urls = ["https://www.harveynorman.ie/index.php?dispatch=xmlsitemap.view"]
    now_time = time.localtime()
    now_date = time.strftime("%d-%m-%Y", now_time)
    time_on = time.strftime("%I:%M %p", now_time)
    timer_start = time.time()
    current_date = datetime.now().strftime('%d_%m_%Y')
    current_timestamp = datetime.now()

#---------------------------------------------------- 
    wb = openpyxl.Workbook()
    # load_workbook = openpyxl.load_workbook('harve_ie_prod.xlsx')
    # load_sheet = load_workbook.active
    data_sheet = wb.active
    data_sheet.title = 'Product_collection'
    data_sheet["A1"] = "domain_url"
    data_sheet["B1"] = "product_url"
    data_sheet["C1"] = "sku"
    data_sheet["D1"] = "mfg_number"
    data_sheet["E1"] = "gtin"
    data_sheet["F1"] = "parent_sku"
    data_sheet["G1"]= "title"
    data_sheet["H1"] = "brand"
    data_sheet["I1"] = "end_category"
    data_sheet["J1"] = "breadcrumbs"
    data_sheet["K1"] = "description"
    data_sheet["L1"] = "features"
    data_sheet["M1"] = "price"
    data_sheet["N1"] = "strike_price"
    data_sheet["O1"] = "uom"
    data_sheet["P1"] = "upc"
    data_sheet["Q1"] = "quantity"
    data_sheet["R1"] = "stock"
    data_sheet["S1"] = "weight"
    data_sheet["T1"] = "warning"
    data_sheet["U1"] = "time_strap"
    data_sheet["V1"] = "industry_id"
    data_sheet["W1"] = "meta_title"
    data_sheet["X1"] = "meta_description"
    data_sheet["Y1"] = "meta_keywords"
    data_sheet["Z1"] = "other_details"
    data_sheet["AA1"] = "ratings"
    data_sheet["AB1"] = "reviews"
    data_sheet["AC1"] = "5_ratings"
    data_sheet["AD1"] = "1_ratings"
    # ---------------------------------------------------- 
    fs2 = wb.create_sheet('Facet')
    fs2['A1'] = 'domain_url'   
    fs2['B1'] = 'category_url'  
    fs2['C1'] = 'category_name'  
    fs2['D1'] = 'facet_name'   
    fs2['E1'] = 'facet_value'
# ---------------------------------------------------- 

# ---------------------------------------------------- 
    ws2 = wb.create_sheet('specification')
    ws2['A1'] = 'domain_url'   
    ws2['B1'] = 'product_url'   
    ws2['C1'] = 'sku'   
    ws2['D1'] = 'attribute_group'
    ws2['E1'] = 'attribute_name'
    ws2['F1'] = 'attribute_value'
# ---------------------------------------------------- 
    media_sheet = wb.create_sheet('Media')
    media_sheet['A1'] = 'domain_url'   
    media_sheet['B1'] = 'product_url'   
    media_sheet['C1'] = 'sku'   
    media_sheet['D1'] = 'media_type'
    media_sheet['E1'] = 'media_url'
    media_sheet['F1'] = 'media_name'
    media_sheet['G1'] = 'media_counts'
   
    max_row = 2
    spec_sheet = 2
    media_sheets =2
    facet_sheet = 2
    temporary_value = 500
    file_path = getcwd()     
    warnings.filterwarnings('ignore')

    def clean(self,text):
        text = re.sub(r'\n+','',text)
        text = re.sub(r'\s+',' ',text)
        text = re.sub(r'\r+',' ',text)
        return text.strip()
    
    def parse(self, response):
        if re.findall(r'<loc>(.*?)<\/loc>',response.text):
            url_collection = re.findall(r'<loc>(.*?)<\/loc>',response.text)
            for i in url_collection:
                if '.html' not in i:
                    yield scrapy.Request(i,callback=self.listing_product,dont_filter=True)
    
    def parse_detail(self,response):
        for category_url in response.xpath('//div[@class="nav-col"]/ul/li/a'):
            url = response.urljoin(category_url.xpath('./@href').get(''))            
            yield response.follow(url,callback=self.listing_product,dont_filter=True)

    def listing_product(self,response):
            if response.xpath('//div[@class="content collapse-single-content"]//div[@class="filter-list attraqt_facet "]//div[@class="accordion-drop"]//input'):
                if response.xpath('//div[@class="product-info"]/a'):
                    bread_crumbs = response.xpath('//div[@class="breadcrumbs-container"]//ul//li//text()').getall()
                    category_name = response.xpath('//div[@class="breadcrumbs-container"]//ul/li/span/text()').get('').strip()
                    for block in response.xpath('//div[@class="content collapse-single-content"]//div[@class="filter-list attraqt_facet "]//div[@class="accordion-drop"]//input'):
                        facet_group = ' '.join([i.strip()for i in block.xpath('./ancestor::div[@class="accordion-drop"]/preceding-sibling::div[@class="filter-heading accordion-title "]//text()').getall()]).strip()
                        facet_name = block.xpath('./@value').get('').strip()
                        if '[' in  facet_name :
                            continue
                        if 'Update' in   facet_name:
                            continue                    
                        self.fs2.cell(row =self.facet_sheet, column =1).value ="harvey-norman.co.uk"
                        self.fs2.cell(row =self.facet_sheet, column =2).value =response.url
                        self.fs2.cell(row =self.facet_sheet, column =3).value =category_name
                        self.fs2.cell(row =self.facet_sheet, column =4).value =facet_group 
                        self.fs2.cell(row =self.facet_sheet, column =5).value =facet_name 
                        self.facet_sheet+=1
                    for i in response.xpath('//div[@class="product-info"]/a'):
                        url = response.urljoin(i.xpath('./@href').get(''))                
                        yield response.follow(url,callback=self.product_block,dont_filter=True)
                    next_page = response.xpath('//a[@aria-label="Next page"]/@href|//link[@rel="next"]/@href').get('')
                    if next_page:
                        yield response.follow(next_page,callback=self.listing_product,dont_filter=True)
                    else:
                        for block in response.xpath('//div[@id="main-content-outer"]//ul/li//a'):
                            url_block = block.xpath('./@href').get('')
                            yield response.follow(url_block,callback=self.listing_product,dont_filter=True)
                else:
                    with open('No_product.txt','a') as f:
                        f.write(str(response.url) +'\n')  
            else:
                with open('brand_collection.txt','a') as f:
                        f.write(str(response.url) +'\n')    
    def product_block(self,response):
        item = {}
    
        response_soup = BeautifulSoup(response.text,'html.parser')
        item['Product_url'] = response.url
        
        title = response.xpath('//h1/text()').get('').strip()
        if response.xpath('//input[@value="Add to cart"]/@value'):
            stock = 'In Stock'
        else:
            stock = 'Out Stock'
        price =''.join([i.strip() for i in response.xpath('//span[@class="price  "]//text()|//span[@class="price-num"]//text()').getall()]).strip()
        strike_price = ' '.join([i.replace('RRP','').strip() for i in response.xpath('//span[@class="price-old"]//text()').getall()]).strip()
        sku = re.findall(r'\'sku\'\:\s\'(.*?)\'\,\s*\'image\'',response.text)[0].strip()
        item['sku'] = sku
        yield item
        mpn = re.findall(r"\'mpn\'\:\s*\"(.*?)\"\,\s*'gtin",response.text)[0].strip().replace('n/a','')
        gtin = re.findall(r"\'gtin\'\:\s*\"(.*?)\"\,\s*'offerPrice",response.text)[0].strip()
        warning = response.xpath('//div[@id="content_description"]//*[contains(text(),"Warning")]/following-sibling::ul/li/text()').get('').strip()
        if  response.xpath('//div[@id="content_description"]//p//text()|//div[@id="content_description"]//h5//text()'):
            descriptio = self.clean(' '.join([i.strip() for i in response.xpath('//div[@id="content_description"]//p//text()|//div[@id="content_description"]//h5//text()').getall()]))
        else:
            descriptio = ''
        
        features = ''
        if response.xpath('//h5[contains(text(),"Features")]/following-sibling::ul/li//text()'):
            features = '\n'.join( [i.strip() for i in response.xpath('//h5[contains(text(),"Features:")]/following-sibling::ul/li//text()').getall()])
        
        
        bread_crumbs_empty = []
        bread_crumbs = response.xpath('//div[@class="breadcrumbs-container"]//ul/li//text()').getall()[:-1]
        for bread in bread_crumbs:
            bread_spliting = bread.replace('\n','').strip()
            if bread_spliting!='':
                bread_crumbs_empty.append(bread_spliting)
        brand = response.xpath('//div[@id="content_features"]//strong[contains(text(),"Brand")]/parent::th/following-sibling::td/text()').get('').strip()
        weight = response.xpath('//div[@id="content_features"]//strong[contains(text(),"Weight")]/parent::th/following-sibling::td/text()').get('')
        meta_title = response.xpath('//meta[@property="og:title"]/@content').get('').strip()
        meta_description = response.xpath('//meta[@name="description"]/@content').get('').strip()
        meta_keywords =response.xpath('//meta[@name="keywords"]/@content').get('').strip()
        other_details = ""
        ratings	=response.xpath('//span[@itemprop="ratingValue"]/text()').get('')
        if ratings:
            rating_count = ratings.strip()
        else:
            rating_count = '0'
        reviews	=response.xpath('//span[@itemprop="reviewCount"]/text()').get('')
        if reviews:
            review_count = reviews.strip()
        else:
            review_count = '0'
        five_ratings =""
        one_ratings =""
        self.data_sheet.cell(row =self.max_row, column =1).value ='harveynorman.ie'
        self.data_sheet.cell(row =self.max_row, column =2).value =response.url         
        self.data_sheet.cell(row =self.max_row, column =3).value =sku
        self.data_sheet.cell(row =self.max_row, column =4).value =''
        self.data_sheet.cell(row =self.max_row, column =5).value =gtin
        self.data_sheet.cell(row =self.max_row, column =6).value =sku
        self.data_sheet.cell(row =self.max_row, column =7).value =title
        self.data_sheet.cell(row =self.max_row, column =8).value =brand
        self.data_sheet.cell(row =self.max_row, column =9).value =bread_crumbs_empty[-1]
        self.data_sheet.cell(row =self.max_row, column =10).value ='|'.join(bread_crumbs_empty)
        self.data_sheet.cell(row =self.max_row, column =11).value = descriptio
        self.data_sheet.cell(row =self.max_row, column =12).value = features
        self.data_sheet.cell(row =self.max_row, column =13).value = price
        self.data_sheet.cell(row =self.max_row, column =14).value = strike_price
        self.data_sheet.cell(row =self.max_row, column =15).value = ""
        self.data_sheet.cell(row =self.max_row, column =16).value = ""
        self.data_sheet.cell(row =self.max_row, column =17).value = ""
        self.data_sheet.cell(row =self.max_row, column =18).value = stock
        self.data_sheet.cell(row =self.max_row, column =19).value = weight
        self.data_sheet.cell(row =self.max_row, column =20).value = warning
        self.data_sheet.cell(row =self.max_row, column =21).value = datetime.now().strftime('%d-%m-%Y %I:%M %p')
        self.data_sheet.cell(row =self.max_row, column =22).value = ""
        self.data_sheet.cell(row =self.max_row, column =23).value = ''
        self.data_sheet.cell(row =self.max_row, column =24).value = meta_description
        self.data_sheet.cell(row =self.max_row, column =25).value = meta_keywords
        self.data_sheet.cell(row =self.max_row, column =26).value = other_details
        self.data_sheet.cell(row =self.max_row, column =27).value = rating_count
        self.data_sheet.cell(row =self.max_row, column =28).value = review_count
        self.data_sheet.cell(row =self.max_row, column =29).value = five_ratings
        self.data_sheet.cell(row =self.max_row, column =30).value = one_ratings
        self.max_row +=1
        if response.xpath('//div[@class="flex-video"]/iframe/@src'):
            video_url = set(response.xpath('//div[@class="flex-video"]/iframe/@src').getall())
            for num, video in enumerate(video_url,1):
                self.media_sheet.cell(row =self.media_sheets, column =1).value ="harveynorman.ie"
                self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                self.media_sheet.cell(row =self.media_sheets, column =3).value =sku
                self.media_sheet.cell(row =self.media_sheets, column =4).value ='Video'
                self.media_sheet.cell(row =self.media_sheets, column =5).value ='https:'+video
                self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                self.media_sheets +=1 

        if response.xpath('//div[@class="cm-image-wrap center"]/a/@href'):
            image=[ i.strip() for i in response.xpath('//div[@class="cm-image-wrap center"]/a/@href').getall()]
            for num, images in enumerate(image,1):
                self.media_sheet.cell(row =self.media_sheets, column =1).value ="harveynorman.ie"
                self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                self.media_sheet.cell(row =self.media_sheets, column =3).value =sku
                self.media_sheet.cell(row =self.media_sheets, column =4).value ='Image'
                self.media_sheet.cell(row =self.media_sheets, column =5).value =images
                self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                
                self.media_sheets +=1
        
        soup = BeautifulSoup(response.text,'html5lib')
        
        table_list = soup.select('#content_features .table-product-features')
        if table_list!=[]:   
            spec_list = []         
            for each_table in table_list:
                value_main = each_table.find_previous_sibling('h6').text.strip()
                tr_list = each_table.select('tr')
                if tr_list != []:
                    for each_tr in tr_list:
                        items = {}
                        th_list = each_tr.select('th')
                        td_list = each_tr.select('td')
                        if th_list != [] and td_list != []:
                            print(th_list[0].text.strip(),td_list[0].text.strip())
                            self.ws2.cell(row =self.spec_sheet, column =1).value ="harvey-norman.co.uk"
                            self.ws2.cell(row =self.spec_sheet, column =2).value =response.url
                            self.ws2.cell(row =self.spec_sheet, column =3).value =sku
                            self.ws2.cell(row =self.spec_sheet, column =4).value =value_main.strip()
                            self.ws2.cell(row =self.spec_sheet, column =5).value =th_list[0].text.strip()
                            self.ws2.cell(row =self.spec_sheet, column =6).value =td_list[0].text.strip()
                            self.spec_sheet +=1 
           
                        
       
        if  self.max_row == self.temporary_value:
            try:
                print('----------------------Saving Excel')
                self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
            except:
                print(('Please Close the file'))
                self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
            self.temporary_value+= 500
    
        try:
            print('----------------------Saving Excel')
            self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}_ie.xlsx") 
        except:
            print(('Please Close the file'))
            self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}_ie.xlsx") 
