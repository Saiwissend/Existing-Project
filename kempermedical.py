import re
import scrapy
import warnings
import openpyxl
from os import getcwd,path
from datetime import datetime
from bs4  import BeautifulSoup
import os, os.path, time, json, pathlib


class KempermedicalsSpider(scrapy.Spider):
    name = "kempermedical"
    now_time = time.localtime()
    now_date = time.strftime("%d-%m-%Y", now_time)
    time_on = time.strftime("%I:%M %p", now_time)
    timer_start = time.time()
    current_date = datetime.now().strftime('%d_%m_%Y')
    current_timestamp = datetime.now()

#---------------------------------------------------- 
    wb = openpyxl.Workbook()
    data_sheet = wb.active
    data_sheet.title = 'Product_collection'
    data_sheet["A1"] = "domain_url"
    data_sheet["B1"] = "product_url"
    data_sheet["C1"] = "sku"
    data_sheet["D1"] = "mfg_number"
    data_sheet["E1"] = "item_number"
    data_sheet["F1"]= "parent_sku"
    data_sheet["G1"] = "title"
    data_sheet["H1"] = "brand"
    data_sheet["I1"] = "end_category"
    data_sheet["J1"] = "breadcrumbs"
    data_sheet["K1"] = "description"
    data_sheet["L1"] = "features"
    data_sheet["M1"] = "price"
    data_sheet["N1"] = "strike_price"
    data_sheet["O1"] = "uom"
    data_sheet["P1"] = "upc"
    data_sheet["Q1"] = "quantity"
    data_sheet["R1"] = "stock"
    data_sheet["S1"] = "weight"
    data_sheet["T1"] = "warning"
    data_sheet["U1"] = "time_strap"
    data_sheet["V1"] = "industry_code"
    # ---------------------------------------------------- 
    fs2 = wb.create_sheet('Facet')
    fs2['A1'] = 'domain_url'   
    fs2['B1'] = 'category_url'  
    fs2['C1'] = 'category_name'  
    fs2['D1'] = 'facet_name'   
    fs2['E1'] = 'facet_value'
# ---------------------------------------------------- 

# ---------------------------------------------------- 
    ws2 = wb.create_sheet('specification')
    ws2['A1'] = 'domain_url'   
    ws2['B1'] = 'product_url'   
    ws2['C1'] = 'sku'   
    ws2['D1'] = 'attribute_group'
    ws2['E1'] = 'attribute_name'
    ws2['F1'] = 'attribute_value'
# ---------------------------------------------------- 
    media_sheet = wb.create_sheet('Media')
    media_sheet['A1'] = 'domain_url'   
    media_sheet['B1'] = 'product_url'   
    media_sheet['C1'] = 'sku'   
    media_sheet['D1'] = 'media_type'
    media_sheet['E1'] = 'media_url'
    media_sheet['F1'] = 'media_name'
    media_sheet['G1'] = 'media_counts'
   
    max_row = 2
    spec_sheet = 2
    media_sheets =2
    facet_sheet = 2
    temporary_value = 500
    file_path = getcwd() 
    warnings.filterwarnings('ignore')
# -------------------------------------------#-------------------------------------------------------------
    start_urls = ["https://kempermedical.com/"]
    def clean(self,text):
        '''remove extra spaces & junk character'''
        text = re.sub(r'\n+','',text)
        text = re.sub(r'\s+',' ',text)
        text = re.sub(r'\r+',' ',text)
        return text.strip()
    def parse(self,response):
        sub_url = ['https://kempermedical.com/collections/quickship-products']                    
        category_url = [response.urljoin(i).strip() for i in response.xpath('//header[@class="desctop-menu-large small-header"]/nav/ul/li/ul/li/ul/li/a/@href').getall()]
        category_url.extend(sub_url)
        for num , block in enumerate(category_url,1):                
            # print(f'=------------------{num},---->{block}')
            yield scrapy.Request(block,callback=self.parse_detail)
    def parse_detail(self,response):
        for url in response.xpath('//div[@class="tt-product-listing row"]//h2/a'):
            product_url_collection = url.xpath('./@href').get('')
            yield response.follow(product_url_collection,callback=self.parse_product_detail)
        next_page = response.xpath('//div[@class="show-more"]/a/@href').get('')
        if next_page:
            yield response.follow(next_page,callback=self.parse_detail,dont_filter=True)
    def parse_product_detail(self,response):
        
        varaints = response.xpath('//div[@itemprop="offers"]/link[@itemprop="url"]/@href').getall()
        for varaint_url in varaints:
            # varaint_url = 'https://kempermedical.com/products/rc-imaging-xarm-weight-bearing-platform?variant=39999140560994'
            yield response.follow(varaint_url,callback=self.parse_detail_product,dont_filter=True)
        else:
            yield response.follow(varaint_url,callback=self.parse_detail_product,dont_filter=True)
    def parse_detail_product(self,response):
        bread = re.findall(r'window\.ap_front_settings\.product_info\.collections\.push\({id:.*?\,\s*title\:\s*`(.*?)\`\}\);',response.text)
        response_soup = BeautifulSoup(response.text,'html.parser')
        bread_crumb = response.xpath('//nav[@aria-label="breadcrumbs"]/a/text()').getall()
        title = response.xpath('//h1/text()').get('').strip()   
        price = response.xpath('//div[@class="tt-product-single-info stickprcolheight-js"]/div[@class="tt-price"]/span[@class="sale-price"]/text()|//div[@class="tt-product-single-info stickprcolheight-js"]/div[@class="tt-price"]/span[@class="new-price"]/text()').get()
        strike_price = response.xpath('//div[@class="tt-product-single-info stickprcolheight-js"]/div[@class="tt-price"]/span[@class="old-price"]/text()').get('').strip()
        sku = response.xpath('//meta[@itemprop="sku"]/@content').get('')
        brand = response.xpath('//meta[@itemprop="brand"]/@content').get('')
        
        description = '\n'.join([i.strip() for i in response.xpath('//div[contains(text(),"DESCRIPTION")]/following-sibling::div//text()').getall()])
        feature_list = []
        for split_spec in response.xpath('//div[contains(text(),"DESCRIPTION")]/following-sibling::div//ul/li'):
            text_extract = ' '.join ([i.strip() for i in split_spec.xpath('.//text()').getall()])
            if ':' in text_extract:
                self.ws2.cell(row =self.spec_sheet, column =1).value ="kempermedical.com"
                self.ws2.cell(row =self.spec_sheet, column =2).value =response.url
                self.ws2.cell(row =self.spec_sheet, column =3).value =sku
                self.ws2.cell(row =self.spec_sheet, column =4).value =''
                self.ws2.cell(row =self.spec_sheet, column =5).value =text_extract.split(':')[0].strip() if ':' in text_extract else ''
                self.ws2.cell(row =self.spec_sheet, column =6).value =text_extract.split(':')[-1].strip() if ':' in text_extract else ''
                self.spec_sheet +=1 
            else:
                feature_list.append(text_extract)
        self.data_sheet.cell(row =self.max_row, column =1).value ='kempermedical.com'
        self.data_sheet.cell(row =self.max_row, column =2).value =response.url         
        self.data_sheet.cell(row =self.max_row, column =3).value =sku
        self.data_sheet.cell(row =self.max_row, column =4).value =""
        self.data_sheet.cell(row =self.max_row, column =5).value =""
        self.data_sheet.cell(row =self.max_row, column =6).value =""
        self.data_sheet.cell(row =self.max_row, column =7).value =title
        self.data_sheet.cell(row =self.max_row, column =8).value =brand
        self.data_sheet.cell(row =self.max_row, column =9).value =bread[-1]
        self.data_sheet.cell(row =self.max_row, column =10).value ='>'.join(bread)
        self.data_sheet.cell(row =self.max_row, column =11).value = description
        self.data_sheet.cell(row =self.max_row, column =12).value = '\n'.join(feature_list if feature_list !=[] else '')
        self.data_sheet.cell(row =self.max_row, column =13).value = price
        self.data_sheet.cell(row =self.max_row, column =14).value = strike_price
        self.data_sheet.cell(row =self.max_row, column =15).value = ""
        self.data_sheet.cell(row =self.max_row, column =16).value = ""
        self.data_sheet.cell(row =self.max_row, column =17).value = ""
        self.data_sheet.cell(row =self.max_row, column =18).value = ""
        self.data_sheet.cell(row =self.max_row, column =19).value = ""
        self.data_sheet.cell(row =self.max_row, column =20).value = ""
        self.data_sheet.cell(row =self.max_row, column =21).value = datetime.now().strftime('%d-%m-%Y %I:%M %p')
        self.data_sheet.cell(row =self.max_row, column =22).value = ""
        self.max_row +=1
        if response.xpath('//ul[@id="smallGallery"]//li//img/@srcset'):
            image=[('https:'+ i).split('&')[0] for i in response.xpath('//ul[@id="smallGallery"]//li//img/@srcset').getall()]
            for num, images in enumerate(image,1):
                self.media_sheet.cell(row =self.media_sheets, column =1).value ="kempermedical.com"
                self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                self.media_sheet.cell(row =self.media_sheets, column =3).value =sku
                self.media_sheet.cell(row =self.media_sheets, column =4).value ='Image'
                self.media_sheet.cell(row =self.media_sheets, column =5).value =images
                self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                self.media_sheets +=1 
        
        if response.xpath('//a[contains(@href,".pdf")]'):
            pdf = set(response.xpath('//a[contains(@href,".pdf")]/@href').getall())
            for num, video in enumerate(pdf,1):
                self.media_sheet.cell(row =self.media_sheets, column =1).value ="kempermedical.com"
                self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                self.media_sheet.cell(row =self.media_sheets, column =3).value =sku
                self.media_sheet.cell(row =self.media_sheets, column =4).value ='PDF'
                self.media_sheet.cell(row =self.media_sheets, column =5).value =video
                self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                self.media_sheets +=1 
        if re.findall(r'<h6>Additional\s*Information<\/h6>([\w\W]*?)<\/div>',response.text):
            addition_info = re.findall(r'<h6>Additional\s*Information<\/h6>([\w\W]*?)<\/div>',response.text)[0].split('<br>')
            for split_spec in addition_info:
                text_extract = split_spec.strip()
                if ':' in text_extract:
                    self.ws2.cell(row =self.spec_sheet, column =1).value ="kempermedical.com"
                    self.ws2.cell(row =self.spec_sheet, column =2).value =response.url
                    self.ws2.cell(row =self.spec_sheet, column =3).value =sku
                    self.ws2.cell(row =self.spec_sheet, column =4).value =''
                    self.ws2.cell(row =self.spec_sheet, column =5).value =text_extract.split(':')[0].strip() if ':' in text_extract else ''
                    self.ws2.cell(row =self.spec_sheet, column =6).value =text_extract.split(':')[-1].strip() if ':' in text_extract else ''
                    self.spec_sheet +=1 
        try:
            print('----------------------Saving Excel')
            self.wb.save(f"{self.file_path}\\Kemper_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
        except:
            print(('Please Close the file'))
            self.wb.save(f"{self.file_path}\\Kemper_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 



