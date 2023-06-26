import re
import scrapy
import warnings
import openpyxl
from os import getcwd,path
from datetime import datetime
from bs4  import BeautifulSoup
import os, os.path, time, json, pathlib
import sqlite3 as sql
import pandas as pd

class McfaddensSpider(scrapy.Spider):
    name = "mcfaddens"
    now_time = time.localtime()
    now_date = time.strftime("%d-%m-%Y", now_time)
    time_on = time.strftime("%I:%M %p", now_time)
    timer_start = time.time()
    current_date = datetime.now().strftime('%d_%m_%Y')
    current_timestamp = datetime.now()
    database = "./assets/facet.db"
#---------------------------------------------------- 
    wb = openpyxl.Workbook()
    data_sheet = wb.active
    data_sheet.title = 'Product_collection'
    data_sheet["A1"] = "domain_url"
    data_sheet["B1"] = "product_url"
    data_sheet["C1"] = "sku"
    data_sheet["D1"] = "mfg_number"
    data_sheet["E1"] = "item_number"
    data_sheet["F1"]= "parent_sku"
    data_sheet["G1"] = "title"
    data_sheet["H1"] = "brand"
    data_sheet["I1"] = "end_category"
    data_sheet["J1"] = "breadcrumbs"
    data_sheet["K1"] = "description"
    data_sheet["L1"] = "features"
    data_sheet["M1"] = "price"
    data_sheet["N1"] = "strike_price"
    data_sheet["O1"] = "uom"
    data_sheet["P1"] = "upc"
    data_sheet["Q1"] = "quantity"
    data_sheet["R1"] = "stock"
    data_sheet["S1"] = "weight"
    data_sheet["T1"] = "warning"
    data_sheet["U1"] = "time_strap"
    data_sheet["V1"] = "industry_code"

# ---------------------------------------------------- 
    ws2 = wb.create_sheet('specification')
    ws2['A1'] = 'domain_url'   
    ws2['B1'] = 'product_url'   
    ws2['C1'] = 'sku'   
    ws2['D1'] = 'attribute_group'
    ws2['E1'] = 'attribute_name'
    ws2['F1'] = 'attribute_value'
# ---------------------------------------------------- 
    media_sheet = wb.create_sheet('Media')
    media_sheet['A1'] = 'domain_url'   
    media_sheet['B1'] = 'product_url'   
    media_sheet['C1'] = 'sku'   
    media_sheet['D1'] = 'media_type'
    media_sheet['E1'] = 'media_url'
    media_sheet['F1'] = 'media_name'
    media_sheet['G1'] = 'media_counts'
# ---------------------------------------------------- 
    fs2 = wb.create_sheet('Facet')
    fs2['A1'] = 'domain_url'   
    fs2['B1'] = 'category_url'   
    fs2['C1'] = 'category_name'   
    fs2['D1'] = 'facet_group'   
    fs2['E1'] = 'facet_value'
# ---------------------------------------------------- 
    max_row = 2
    spec_sheet = 2
    media_sheets =2
    temporary_value = 500
    facet_sheet=2
    file_path = getcwd() 
# -------------------------------------------#-------------------------------------------------------------
    start_urls = ["https://mcfaddens.com/"]
    def clean(self,text):
        '''remove extra spaces & junk character'''
        text = re.sub(r'\n+','',text)
        text = re.sub(r'\s+',' ',text)
        text = re.sub(r'\r+',' ',text)
        return text.strip()
    def parse(self,response):
        for block in response.xpath('//div[contains(text(),"Categories")]/parent::div//a'):
            url_categories = block.xpath('./@href').get('')
            yield response.follow(url_categories,callback=self.parse_detail,dont_filter=True)
    def parse_detail(self,response):
        if response.xpath('//div[@id="center-main"]//span[@class="subcategories"]/div[@class="subcategory-content"]/a'):
            for block in response.xpath('//div[@id="center-main"]//span[@class="subcategories"]/div[@class="subcategory-content"]/a'):
                url_categories = block.xpath('./@href').get('')
                yield response.follow(url_categories,callback=self.parse_cate,dont_filter=True)
        else:
            yield response.follow(response.url,callback=self.parse_detail)
    def parse_cate(self,response): 
# --------------------------------------FACET_SHEET-----#-------------------------------------------------------------       
        category_name = response.xpath('//div[@class="common-header"]/text()').get('')       
        for block in response.xpath('//div[@class="rf-customer-filters"]//div[@class="rf-filters-buttons"]/following-sibling::table/tr//span[@class="rf-txt-active"]/label'):
            facet_group = ' '.join([i.strip()for i in block.xpath('./ancestor::div[@class="rf-element-box"]/preceding-sibling::table//span[@class="rf-element-title red"]/text()').getall()]).strip()
            
            facet_name = block.xpath('./text()').get('').strip()
        
            self.fs2.cell(row =self.facet_sheet, column =1).value ="https://mcfaddens.com/"
            self.fs2.cell(row =self.facet_sheet, column =2).value =response.url 
            self.fs2.cell(row =self.facet_sheet, column =3).value =category_name 
            self.fs2.cell(row =self.facet_sheet, column =4).value =facet_group 
            self.fs2.cell(row =self.facet_sheet, column =5).value =facet_name 
            self.facet_sheet+=1
        if response.xpath('//span[@class="rf-element-title red"][contains(text(),"Price")]'):
             for i in response.xpath('//label[contains(text(),"Min Value")]/following-sibling::input/@value|//label[contains(text(),"Max Value")]/following-sibling::input/@value').getall():
                facet_name = i.strip() 
                facet_group = 'Price'
                self.fs2.cell(row =self.facet_sheet, column =1).value ="https://mcfaddens.com/"
                self.fs2.cell(row =self.facet_sheet, column =2).value =response.url 
                self.fs2.cell(row =self.facet_sheet, column =3).value =category_name 
                self.fs2.cell(row =self.facet_sheet, column =4).value =facet_group 
                self.fs2.cell(row =self.facet_sheet, column =5).value =facet_name 
                self.facet_sheet+=1
# -------------------------------------------#-------------------------------------------------------------
        for block in response.xpath('//div[@class="image"]/a'):
            url_categories = block.xpath('./@href').get('')
            yield response.follow(url_categories,callback=self.parse_product,dont_filter=True)
        next_page = response.xpath('//div[@class="nav-pages"]//a[@class="right-arrow nav-navigation-link"]/@href').get('')
        if next_page:
            yield response.follow(next_page,callback=self.parse_cate,dont_filter=True)
    def parse_product(self,response):
        try:
            item = {}
            response_soup = BeautifulSoup(response.text,'html.parser')
            source_url = response.url
            title = response.xpath('//h1[@class="product-main-title"]/text()').get('').strip()
            manufacturer = response.xpath("//td[contains(text(),'Manufacturer')]/following-sibling::td/text()").get('').strip()
            if re.findall(r'\<tr\>\s*<td class=\"property\-feature\-name\"\>Manufacturer\'s Part\<\/td\>\s*<td class\=\"property-value\"\>(.*?)\<\/td>\s*<\/tr>',response.text):
                mpn =re.findall(r'\<tr\>\s*<td class=\"property\-feature\-name\"\>Manufacturer\'s Part\<\/td\>\s*<td class\=\"property-value\"\>(.*?)\<\/td>\s*<\/tr>',response.text)[0]
            else:
                mpn = ''
            part_Number = response.xpath("//td[contains(text(),'Part Number')]/following-sibling::td/text()").get('').strip()
            price = ''.join(response.xpath('//span[@class="product-price-value"]//text()').getall()).strip()
            strike_price = response.xpath('//span[@class="product-market-price"]//text()').get('').strip()
            old_price = ''
            if price ==strike_price:
                pass
            else:
                old_price = strike_price
# -----------------------------------------Image Sheet --#-------------------------------------------------------------            
            image = []
            if response.xpath('//div[@data-type="image-additional"]//a/@href|//img[@id="product_thumbnail"]/@src'):
                image=set(response.xpath('//div[@data-type="image-additional"]//a/@href').getall())
                for num, images in enumerate(image,1):
                        self.media_sheet.cell(row =self.media_sheets, column =1).value ="mcfaddens.com"
                        self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                        self.media_sheet.cell(row =self.media_sheets, column =3).value =part_Number
                        self.media_sheet.cell(row =self.media_sheets, column =4).value ='Image'
                        self.media_sheet.cell(row =self.media_sheets, column =5).value =images
                        self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                        self.media_sheet.cell(row =self.media_sheets, column =7).value =num
            
            bread_crumb_list = response.xpath('//div[@id="location"]//span[@itemprop="name"]//text()').getall()[0:-1]
            descriptio = '\n'.join([i.strip() for i in response.xpath('//span[@itemprop="description"]//text()').getall()])
            stock = response.xpath('//*[contains(text(),"In Stock")]/text()').get('').strip()
            weight = ' '.join([i.strip() for i in response.xpath('//div[@id="block_product_details"]//div[@data-type="features"]//td[contains(text(),"Weight")]/following-sibling::td//text()').getall()])
# -------------------------------------------#-Video ------------------------------------------------------------
            empty_url = []
            if response.xpath('//h2[contains(text(),"Additional resources")]/parent::div/following-sibling::div//a[contains(@onclick,"javascript")]/@href'):
                media_url = response.xpath('//h2[contains(text(),"Additional resources")]/parent::div/following-sibling::div//a[contains(@onclick,"javascript")]/@onclick').get('')
                if re.findall(r'javascript.*?\/embed\/(.*?)&filename.*?',media_url):
                    for i in re.findall(r'javascript.*?\/embed\/(.*?)&filename.*?',media_url):
                        youtube_url = 'https://www.youtube.com/watch?v='+ i
                        empty_url.append(youtube_url)
                    for num, videos in enumerate(empty_url,1):
                        self.media_sheet.cell(row =self.media_sheets, column =1).value ="mcfaddens.com"
                        self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                        self.media_sheet.cell(row =self.media_sheets, column =3).value =part_Number
                        self.media_sheet.cell(row =self.media_sheets, column =4).value ='Video'
                        self.media_sheet.cell(row =self.media_sheets, column =5).value =videos
                        self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                        self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                        self.media_sheets+=1
                else:
                    youtube_url = []
# -------------------------------------------#-PDF ------------------------------------------------------------
            pdf_url = ['https:'+i for i in response.xpath('//h2[contains(text(),"Additional resources")]/parent::div/following-sibling::div//a[contains(@href,".pdf")]/@href').getall()]
            for num, pdf in enumerate(pdf_url,1):
                        self.media_sheet.cell(row =self.media_sheets, column =1).value ="mcfaddens.com"
                        self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                        self.media_sheet.cell(row =self.media_sheets, column =3).value =part_Number
                        self.media_sheet.cell(row =self.media_sheets, column =4).value ='PDF'
                        self.media_sheet.cell(row =self.media_sheets, column =5).value =pdf
                        self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                        self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                        self.media_sheets+=1
# -------------------------------------------#-------------------------------------------------------------
            self.data_sheet.cell(row =self.max_row, column =1).value ='mcfaddens.com'
            self.data_sheet.cell(row =self.max_row, column =2).value =response.url         
            self.data_sheet.cell(row =self.max_row, column =3).value =part_Number
            self.data_sheet.cell(row =self.max_row, column =4).value =mpn
            self.data_sheet.cell(row =self.max_row, column =5).value =""
            self.data_sheet.cell(row =self.max_row, column =6).value =""
            self.data_sheet.cell(row =self.max_row, column =7).value =title
            self.data_sheet.cell(row =self.max_row, column =8).value =manufacturer
            self.data_sheet.cell(row =self.max_row, column =9).value =bread_crumb_list[-1]
            self.data_sheet.cell(row =self.max_row, column =10).value ='>'.join(bread_crumb_list)
            self.data_sheet.cell(row =self.max_row, column =11).value = descriptio
            self.data_sheet.cell(row =self.max_row, column =12).value = ""
            self.data_sheet.cell(row =self.max_row, column =13).value = price
            self.data_sheet.cell(row =self.max_row, column =14).value = old_price
            self.data_sheet.cell(row =self.max_row, column =15).value = ""
            self.data_sheet.cell(row =self.max_row, column =16).value = ""
            self.data_sheet.cell(row =self.max_row, column =17).value = ""
            self.data_sheet.cell(row =self.max_row, column =18).value = stock
            self.data_sheet.cell(row =self.max_row, column =19).value = weight
            self.data_sheet.cell(row =self.max_row, column =20).value = ""
            self.data_sheet.cell(row =self.max_row, column =21).value = datetime.now().strftime('%d-%m-%Y %I:%M %p')
            self.data_sheet.cell(row =self.max_row, column =22).value = ""
            self.max_row +=1
# -------------------------------------------Specification Sheet -------------------------------------------------------------            
            for block in response.xpath('//div[@id="block_product_details"]//div[@data-type="features"]//tr'):
                assing_key = block.xpath('./td[1]/text()').get('').strip()
                assing_value = ' '.join([i.strip() for i in block.xpath('./td[2]//text()').getall()])
                self.ws2.cell(row =self.spec_sheet, column =1).value ="harvey-norman.co.uk"
                self.ws2.cell(row =self.spec_sheet, column =2).value =response.url
                self.ws2.cell(row =self.spec_sheet, column =3).value =part_Number
                self.ws2.cell(row =self.spec_sheet, column =4).value =''
                self.ws2.cell(row =self.spec_sheet, column =5).value =assing_key
                self.ws2.cell(row =self.spec_sheet, column =6).value =assing_value
                self.spec_sheet +=1 
            if  self.max_row == self.temporary_value:
                try:
                    print('----------------------Saving Excel')
                    self.wb.save(f"{self.file_path}\\mcfadeens_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
                except:
                    print(('Please Close the file'))
                    self.wb.save(f"{self.file_path}\\mcfadeens_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
                self.temporary_value+= 500  
        # try:
        #         self.wb.save(f"{file_path}\\mcfadeens_25_05_2023.xlsx") 
        except Exception as e:
            print(e)
            with open('issue.txt','a') as f:
                        f.write(str(response.url) +'\n') 
        try:
            print('----------------------Saving Excel')
            self.wb.save(f"{self.file_path}\\mcfadeens_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
        except:
            print(('Please Close the file'))
            self.wb.save(f"{self.file_path}\\mcfadeens_{datetime.now().strftime('%d_%m_%Y')}.xlsx")  
        # yield item
        