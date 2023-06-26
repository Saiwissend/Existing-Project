import re
import scrapy
import warnings
import openpyxl
from os import getcwd,path
from datetime import datetime
from bs4  import BeautifulSoup
import os, os.path, time, json, pathlib
from datetime import datetime
import sqlite3 as sql
import pandas as pd



class harveySpider(scrapy.Spider):
    name = "harvey"
    now_time = time.localtime()
    now_date = time.strftime("%d-%m-%Y", now_time)
    time_on = time.strftime("%I:%M %p", now_time)
    timer_start = time.time()
    current_date = datetime.now().strftime('%d_%m_%Y')
    current_timestamp = datetime.now()

#---------------------------------------------------- 
    wb = openpyxl.Workbook()    
    data_sheet = wb.active
    data_sheet.title = 'Product_collection'
    data_sheet["A1"] = "domain_url"
    data_sheet["B1"] = "product_url"
    data_sheet["C1"] = "sku"
    data_sheet["D1"] = "mfg_number"
    data_sheet["E1"] = "item_number"
    data_sheet["F1"]= "parent_sku"
    data_sheet["G1"] = "title"
    data_sheet["H1"] = "brand"
    data_sheet["I1"] = "end_category"
    data_sheet["J1"] = "breadcrumbs"
    data_sheet["K1"] = "description"
    data_sheet["L1"] = "features"
    data_sheet["M1"] = "price"
    data_sheet["N1"] = "strike_price"
    data_sheet["O1"] = "uom"
    data_sheet["P1"] = "upc"
    data_sheet["Q1"] = "quantity"
    data_sheet["R1"] = "stock"
    data_sheet["S1"] = "weight"
    data_sheet["T1"] = "warning"
    data_sheet["U1"] = "time_strap"
    data_sheet["V1"] = "industry_id"
    data_sheet["W1"] = "meta_title"
    data_sheet["X1"] = "meta_description"
    data_sheet["Y1"] = "meta_keywords"
    data_sheet["Z1"] = "other_details"
    data_sheet["AA1"] = "ratings"
    data_sheet["AB1"] = "reviews"
    data_sheet["AC1"] = "5_ratings"
    data_sheet["AC1"] = "1_ratings"
    # ---------------------------------------------------- 
    fs2 = wb.create_sheet('Facet')
    fs2['A1'] = 'domain_url'   
    fs2['B1'] = 'category_url'  
    fs2['C1'] = 'category_name'  
    fs2['D1'] = 'facet_name'   
    fs2['E1'] = 'facet_value'
# ---------------------------------------------------- 

# ---------------------------------------------------- 
    ws2 = wb.create_sheet('specification')
    ws2['A1'] = 'domain_url'   
    ws2['B1'] = 'product_url'   
    ws2['C1'] = 'sku'   
    ws2['D1'] = 'attribute_group'
    ws2['E1'] = 'attribute_name'
    ws2['F1'] = 'attribute_value'
# ---------------------------------------------------- 
    media_sheet = wb.create_sheet('Media')
    media_sheet['A1'] = 'domain_url'   
    media_sheet['B1'] = 'product_url'   
    media_sheet['C1'] = 'sku'   
    media_sheet['D1'] = 'media_type'
    media_sheet['E1'] = 'media_url'
    media_sheet['F1'] = 'media_name'
    media_sheet['G1'] = 'media_counts'
   
    max_row = 2
    spec_sheet = 2
    media_sheets =2
    facet_sheet = 2
    temporary_value = 500
    file_path = getcwd() 
    # warnings.filterwarnings('ignore')
# ----------------------------------------------------

    start_urls = ["https://www.harvey-norman.co.uk/"]

    def clean(self,text):
        '''remove extra spaces & junk character'''
        text = re.sub(r'\n+','',text)
        text = re.sub(r'\s+',' ',text)
        text = re.sub(r'\r+',' ',text)
        return text.strip()
    
    def parse(self,response):
        url_collection = []
        for category_url in response.xpath('//div[@class="flex items-center space-x-2"]/following-sibling::ul/li/a'):
            url = response.urljoin(category_url.xpath('./@href').get(''))
            text = (category_url.xpath('./text()').get(''))
            url_collection.append(url)
            yield response.follow(url,callback=self.listing_product,dont_filter=True)

    def listing_product(self,response):
            if response.xpath('//div[@class="card-wrapper hn-card-wrapper"]//div[@class="card card--product card--outline"]/parent::a'):
                category_name = response.xpath('//h1/text()').get('').strip()
                for block in response.xpath('//div[@class="accordion facet-filter-elements"]//div[contains(@class,"accordion-item")]//div[@class="accordion-body"]//input'):
                    facet_group = ' '.join([i.strip()for i in block.xpath('./ancestor::div[@class="accordion-collapse collapse show"]/preceding-sibling::h2//text()').getall()]).strip()
                    facet_name = block.xpath('./@value').get('').strip()
                    if facet_name =='':
                        if block.css('.product-price-input__box.pricerange-min'):                            
                            facet_name = '0.00'
                        if block.css('.product-price-input__box.pricerange-max'):
                            facet_name = block.xpath('./ancestor::hn-price-range-filter/@data-maxprice').get('').strip() 
                    self.fs2.cell(row =self.facet_sheet, column =1).value ="harvey-norman.co.uk"
                    self.fs2.cell(row =self.facet_sheet, column =2).value =response.url
                    self.fs2.cell(row =self.facet_sheet, column =3).value =category_name
                    self.fs2.cell(row =self.facet_sheet, column =4).value =facet_group 
                    self.fs2.cell(row =self.facet_sheet, column =5).value =facet_name 
                    self.facet_sheet+=1
                
                bread_crumb_list = []
                bread_crumbs = response.xpath('//nav[@class="breadcrumb"]//text()').getall()
                for bread in bread_crumbs:
                    bread_spliting = bread.replace('\n','').strip()
                    if bread_spliting!='':
                        bread_crumb_list.append(bread_spliting)
                for i in response.xpath('//div[@class="card-wrapper hn-card-wrapper"]//div[@class="card card--product card--outline"]/parent::a'):
                    url = i.xpath('./@href').get('')        
                    
                    yield response.follow(url,callback=self.product_block,dont_filter=True,cb_kwargs={'bread_crumb_list':bread_crumb_list})
                next_page = response.xpath('//a[@aria-label="Next page"]/@href|//link[@rel="next"]/@href').get('')
                if next_page:
                    yield response.follow(next_page,callback=self.listing_product,dont_filter=True)
            else:
                for block in response.xpath('//div[@id="main-content-outer"]//ul/li//a'):
                    url_block = block.xpath('./@href').get('')
                    yield response.follow(url_block,callback=self.listing_product)
    def product_block(self,response,bread_crumb_list):
        try:
            item = {}
            item['source_url'] = response.url
            response_soup = BeautifulSoup(response.text,'html.parser')
            title = response.xpath('//div[@class="hn-product__title "]/text()').get('').strip()
            if re.findall(r'"availability":.*?org\/\s*(.*?)\s*\"\,',response.text):
                stock = re.findall(r'"availability":.*?org\/\s*(.*?)\s*\"\,',response.text)[0]
            else:
                stock = 'Out Stock'
                
            price = ' '.join([i.strip() for i in response.xpath('//div[@class="price"]/text()').getall()])
            strike_price = ' '.join([i.replace('RRP','').strip() for i in response.xpath('//span[@class="price-rrp"]/text()').getall()]).strip()
            sku = response.xpath('//span[@class="variant-sku variant-sku__font"]/text()').get('').strip()
            if  response_soup.select('.hn-tab-list__description p'):
                descriptio = ' '.join([i.text.strip() for i in response_soup.select('.hn-tab-list__description')[0].select('p')])
            else:
                descriptio = ''
            
            features = ''
            if response_soup.select('.hn-tab-list__description ul'):
                features = '\n'.join( [i.text.strip() for i in response_soup.select('.hn-tab-list__description ul')[0].select('li')])
            
            bread_crumbs_empty = []
            if bread_crumb_list==[]:
                bread_crumbs = response.xpath('//nav[@class="breadcrumb"]//text()').getall()
                for bread in bread_crumbs:
                    bread_spliting = bread.replace('\n','').strip()
                    if bread_spliting!='':
                        bread_crumbs_empty.append(bread_spliting)
                        bread_crumb_list = bread_crumbs_empty
            brand = response.xpath('//div[@class="hn-tab-list__spec"]//li//span[contains(text(),"Brand")]/following-sibling::span/text()').get('').strip()
            weight = response.xpath('//div[@class="hn-tab-list__spec"]//li//span[contains(text(),"Weight")]/following-sibling::span/text()').get('')
            meta_title = response.xpath('//meta[@property="og:title"]/@content').get('')
            meta_description = response.xpath('//meta[@property="og:description"]/@content').get('')
            meta_keywords = ""
            other_details = ""
            ratings	=""
            reviews	=""
            five_ratings =""
            one_ratings =""
            self.data_sheet.cell(row =self.max_row, column =1).value ='harvey-norman.co.uk'
            self.data_sheet.cell(row =self.max_row, column =2).value =response.url         
            self.data_sheet.cell(row =self.max_row, column =3).value =sku
            self.data_sheet.cell(row =self.max_row, column =4).value =""
            self.data_sheet.cell(row =self.max_row, column =5).value =""
            self.data_sheet.cell(row =self.max_row, column =6).value =""
            self.data_sheet.cell(row =self.max_row, column =7).value =title
            self.data_sheet.cell(row =self.max_row, column =8).value =brand
            self.data_sheet.cell(row =self.max_row, column =9).value =bread_crumb_list[-1]
            self.data_sheet.cell(row =self.max_row, column =10).value ='|'.join(bread_crumb_list)
            self.data_sheet.cell(row =self.max_row, column =11).value = descriptio
            self.data_sheet.cell(row =self.max_row, column =12).value = features
            self.data_sheet.cell(row =self.max_row, column =13).value = price
            self.data_sheet.cell(row =self.max_row, column =14).value = strike_price
            self.data_sheet.cell(row =self.max_row, column =15).value = ""
            self.data_sheet.cell(row =self.max_row, column =16).value = ""
            self.data_sheet.cell(row =self.max_row, column =17).value = ""
            self.data_sheet.cell(row =self.max_row, column =18).value = stock
            self.data_sheet.cell(row =self.max_row, column =19).value = weight
            self.data_sheet.cell(row =self.max_row, column =20).value = ""
            self.data_sheet.cell(row =self.max_row, column =21).value = datetime.now().strftime('%d-%m-%Y %I:%M %p')
            self.data_sheet.cell(row =self.max_row, column =22).value = ""
            self.data_sheet.cell(row =self.max_row, column =23).value = meta_title
            self.data_sheet.cell(row =self.max_row, column =24).value = meta_description
            self.data_sheet.cell(row =self.max_row, column =25).value = meta_keywords
            self.data_sheet.cell(row =self.max_row, column =26).value = other_details
            self.data_sheet.cell(row =self.max_row, column =27).value = ratings
            self.data_sheet.cell(row =self.max_row, column =28).value = reviews
            self.data_sheet.cell(row =self.max_row, column =29).value = five_ratings
            self.data_sheet.cell(row =self.max_row, column =30).value = one_ratings
            self.max_row +=1
            if response.xpath('//div[@class="hn-desktop"]//div[@class="flex-video"]/iframe/@src'):
                video_url = set(response.xpath('//div[@class="hn-desktop"]//div[@class="flex-video"]/iframe/@src').getall())
                for num, video in enumerate(video_url,1):
                    self.media_sheet.cell(row =self.media_sheets, column =1).value ="harvey-norman.co.uk"
                    self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                    self.media_sheet.cell(row =self.media_sheets, column =3).value =sku
                    self.media_sheet.cell(row =self.media_sheets, column =4).value ='Video'
                    self.media_sheet.cell(row =self.media_sheets, column =5).value ='https:'+video
                    self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                    self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                    self.media_sheets +=1 

            if response.xpath('//div[@class="product-image"]/div[@id="image-gallery"]/a/@href'):
                image=[('https:'+ i).split('&')[0] for i in response.xpath('//div[@class="product-image"]/div[@id="image-gallery"]/a/@href').getall()]
                for num, images in enumerate(image,1):
                    self.media_sheet.cell(row =self.media_sheets, column =1).value ="harvey-norman.co.uk"
                    self.media_sheet.cell(row =self.media_sheets, column =2).value =response.url
                    self.media_sheet.cell(row =self.media_sheets, column =3).value =sku
                    self.media_sheet.cell(row =self.media_sheets, column =4).value ='Image'
                    self.media_sheet.cell(row =self.media_sheets, column =5).value =images
                    self.media_sheet.cell(row =self.media_sheets, column =6).value =''
                    self.media_sheet.cell(row =self.media_sheets, column =7).value =num
                    
                    self.media_sheets +=1 
            sepc_loop = response_soup.select('.product-image .hn-tab-list__spec')
            if sepc_loop!=[]:
                spec_list = []
                for spec in sepc_loop[0].select('.mb-5'):
                    value_main = spec.select('h4')[0].text
                    for spec_key in spec.select('li'):
                        items = {}
                        items['key'] = spec_key.select('.font-bold.text-base')[0].text.strip()
                        items['value'] = spec_key.select('.w-3\/4')[0].text.strip()
                        items['value_text'] = value_main.strip()
                        spec_list.append(items)
                for spec_loops in spec_list:
                    self.ws2.cell(row =self.spec_sheet, column =1).value ="harvey-norman.co.uk"
                    self.ws2.cell(row =self.spec_sheet, column =2).value =response.url
                    self.ws2.cell(row =self.spec_sheet, column =3).value =sku
                    self.ws2.cell(row =self.spec_sheet, column =4).value =spec_loops.get('value_text')
                    self.ws2.cell(row =self.spec_sheet, column =5).value =spec_loops.get('key')
                    self.ws2.cell(row =self.spec_sheet, column =6).value =spec_loops.get('value')
                    self.spec_sheet +=1 
            if  self.max_row == self.temporary_value:
                try:
                    print('----------------------Saving Excel')
                    self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
                except:
                    print(('Please Close the file'))
                    self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
                self.temporary_value+= 500
            yield item
        except Exception as e:
            print(e)
            with open('issue.txt','a') as f:
                        f.write(str(response.url) +'\n')    
        try:
            print('----------------------Saving Excel')
            self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
        except:
            print(('Please Close the file'))
            self.wb.save(f"{self.file_path}\\harvey_{datetime.now().strftime('%d_%m_%Y')}.xlsx") 
